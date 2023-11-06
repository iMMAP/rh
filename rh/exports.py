from django.http import JsonResponse

from django.views import View
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle
from io import BytesIO
import base64
from django.utils import timezone

from .models import Project

#############################################
############### Export Views #################
#############################################

# Define the header style
header_style = NamedStyle(name="header")
header_style.font = Font(bold=True)


class ProjectExportExcelView(View):
    """
    View class for exporting project data to Excel.
    """

    def post(self, request, project_id):
        """
        Handle POST request to export project data to Excel.

        Args:
            request (HttpRequest): The HTTP request object.
            project_id (int): The ID of the project.

        Returns:
            JsonResponse: The JSON response containing the file URL and name, or an error message.
        """
        try:
            project = Project.objects.get(id=project_id)  # Get the project object

            workbook = Workbook()

            self.write_project_sheet(workbook, project)
            self.write_population_sheet(workbook, project)
            self.write_target_locations_sheet(workbook, project)
            self.write_budget_progress_sheet(workbook, project)

            excel_file = BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)

            response = {
                "file_url": "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
                + base64.b64encode(excel_file.read()).decode("utf-8"),
                "file_name": "export.xlsx",
            }

            return JsonResponse(response)
        except Exception as e:
            response = {"error": str(e)}
            return JsonResponse(response, status=500)

    def write_project_sheet(self, workbook, project):
        """
        Write the project sheet to the workbook.

        Args:
            workbook (Workbook): The Excel workbook object.
            project (Project): The project object.
        """
        sheet = workbook.active
        sheet.title = "Project"

        # Define column headers and types
        columns = [
            {"header": "Focal Person", "type": "string", "width": 20},
            {"header": "Email", "type": "string", "width": 25},
            {"header": "Project Title", "type": "string", "width": 40},
            {"header": "Code", "type": "string", "width": 20},
            {"header": "Project Description", "type": "string", "width": 50},
            {"header": "HRP Project Code", "type": "string", "width": 20},
            {"header": "Project Start Date", "type": "date", "width": 20},
            {"header": "Project End Date", "type": "date", "width": 20},
            {"header": "Project Budget", "type": "float", "width": 10},
            {"header": "Budget Received", "type": "float", "width": 10},
            {"header": "Budget Gap", "type": "float", "width": 10},
            {"header": "Project Budget Currency", "type": "string", "width": 10},
            {"header": "Project Donors", "type": "string", "width": 30},
            {"header": "Implementing Partners", "type": "string", "width": 30},
            {"header": "Programme Partners", "type": "string", "width": 30},
            {"header": "Status", "type": "string", "width": 10},
            {"header": "URL", "type": "string", "width": 20},
        ]

        self.write_sheet_columns(sheet, columns)
        self.write_project_data_rows(sheet, project)

    def write_sheet_columns(self, sheet, columns):
        """
        Write column headers and formatting for a sheet.

        Args:
            sheet (Worksheet): The worksheet object.
            columns (list): List of dictionaries containing column header, type, and width information.
        """
        for idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=idx, value=column["header"])
            cell.style = header_style

            column_letter = get_column_letter(idx)
            if column["type"] == "number":
                sheet.column_dimensions[column_letter].number_format = "General"
            elif column["type"] == "date":
                sheet.column_dimensions[column_letter].number_format = "mm-dd-yyyy"

            sheet.column_dimensions[column_letter].width = column["width"]

    def write_project_data_rows(self, sheet, project):
        """
        Write project data rows to the sheet.

        Args:
            sheet (Worksheet): The worksheet object.
            project (Project): The project object.
        """
        row = [
            project.user.profile.name,
            project.user.email,
            project.title,
            project.code,
            project.description,
            project.hrp_code,
            project.start_date.astimezone(timezone.utc).replace(tzinfo=None) if project.start_date else None,
            project.end_date.astimezone(timezone.utc).replace(tzinfo=None) if project.end_date else None,
            project.budget,
            project.budget_received,
            project.budget_gap,
            project.budget_currency.name,
            ", ".join([donor.name for donor in project.donors.all()]),
            ", ".join([implementing_partner.name for implementing_partner in project.implementing_partners.all()]),
            ", ".join([programme_partner.name for programme_partner in project.programme_partners.all()]),
            project.state,
            "URL",
        ]

        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=2, column=col_idx, value=value)

        sheet.freeze_panes = sheet["A2"]

    def write_population_sheet(self, workbook, project):
        """
        Write the population sheet to the workbook.

        Args:
            workbook (Workbook): The Excel workbook object.
            project (Project): The project object.
        """
        sheet = workbook.create_sheet(title="Target Population")

        # Define column headers and types for Sheet 2
        columns = [
            {"header": "Activity Domain", "type": "string", "width": 20},
            {"header": "Activity Type", "type": "string", "width": 20},
            {"header": "Activity Detail", "type": "string", "width": 20},
            {"header": "Indicators", "type": "string", "width": 40},
        ]

        self.write_sheet_columns(sheet, columns)
        self.write_population_data_rows(sheet, project)

        sheet.freeze_panes = sheet["A2"]

    def write_population_data_rows(self, sheet, project):
        """
        Write population data rows to the sheet.

        Args:
            sheet (Worksheet): The worksheet object.
            project (Project): The project object.
        """
        activity_plans = project.activityplan_set.all()
        for plan in activity_plans:
            row = [
                plan.activity_domain.name,
                plan.activity_type.name,
                plan.activity_detail.name if plan.activity_detail else None,
                "\n".join([indicator.name for indicator in plan.indicators.all()]),
            ]

            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=2, column=col_idx, value=value)

    def write_target_locations_sheet(self, workbook, project):
        """
        Write the target locations sheet to the workbook.

        Args:
            workbook (Workbook): The Excel workbook object.
            project (Project): The project object.
        """
        sheet = workbook.create_sheet(title="Target Locations")

        # Define column headers and types for Sheet 3
        columns = [
            {"header": "Province", "type": "string", "width": 20},
            {"header": "District", "type": "string", "width": 20},
            {"header": "Zone/Ward", "type": "string", "width": 20},
        ]

        self.write_sheet_columns(sheet, columns)
        self.write_target_locations_data_rows(sheet, project)

        sheet.freeze_panes = sheet["A2"]

    def write_target_locations_data_rows(self, sheet, project):
        """
        Write target locations data rows to the sheet.

        Args:
            sheet (Worksheet): The worksheet object.
            project (Project): The project object.
        """
        target_locations = project.targetlocation_set.all()
        for location in target_locations:
            row = [
                location.province.name,
                location.district.name,
                location.zone.name if location.zone else None,
            ]

            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=2, column=col_idx, value=value)

    def write_budget_progress_sheet(self, workbook, project):
        """
        Write the target locations sheet to the workbook.

        Args:
            workbook (Workbook): The Excel workbook object.
            project (Project): The project object.
        """
        sheet = workbook.create_sheet(title="Budget Progress")

        # Define column headers and types for Sheet 3
        columns = [
            {"header": "Project", "type": "string", "width": 20},
            {"header": "Title", "type": "string", "width": 20},
            {"header": "Donor", "type": "string", "width": 20},
            {"header": "Activity Domain", "type": "string", "width": 20},
            {"header": "Grant", "type": "float", "width": 10},
            {"header": "Amount Recieved", "type": "float", "width": 10},
            {"header": "Budget Currency", "type": "string", "width": 20},
            {"header": "Received Date", "type": "date", "width": 20},
            {"header": "Country", "type": "string", "width": 20},
            {"header": "Description", "type": "string", "width": 20},
        ]

        self.write_sheet_columns(sheet, columns)
        self.write_budget_progress_data_rows(sheet, project)

        sheet.freeze_panes = sheet["A2"]

    def write_budget_progress_data_rows(self, sheet, project):
        """
        Write budget progress data rows to the sheet.

        Args:
            sheet (Worksheet): The worksheet object.
            project (Project): The project object.
        """
        budget_progress = project.budgetprogress_set.all()
        for budget in budget_progress:
            row = [
                budget.project.name,
                budget.title,
                budget.donor.name,
                budget.activity_domain.name,
                budget.grant,
                budget.amount_recieved,
                budget.budget_currency.name,
                budget.received_date,
                budget.country.name,
                budget.description,
            ]

            for col_idx, value in enumerate(row, start=1):
                sheet.cell(row=2, column=col_idx, value=value)
