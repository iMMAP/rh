import base64
import csv
import json
from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter

from .models import Disaggregation, Project

#############################################
############### Export Views #################
#############################################

# Define the header style
header_style = NamedStyle(name="header")
header_style.font = Font(bold=True)


class ProjectExportExcelView(View):
    """
    View class for exporting project data to Excel.
    """

    def post(self, request, project_id):
        """
        Handle POST request to export project data to Excel.

        Args:
            request (HttpRequest): The HTTP request object.
            project_id (int): The ID of the project.

        Returns:
            JsonResponse: The JSON response containing the file URL and name, or an error message.
        """
        try:
            project = Project.objects.get(id=project_id)  # Get the project object
            workbook = Workbook()

            self.write_project_sheet(workbook, project)

            excel_file = BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)

            response = {
                "file_url": "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
                + base64.b64encode(excel_file.read()).decode("utf-8"),
                "file_name": "export.xlsx",
            }
            return JsonResponse(response)
        except Exception as e:
            response = {"error": str(e)}
            return JsonResponse(response, status=500)

    def write_project_sheet(self, workbook, project):
        """
        Write the project sheet to the workbook.

        Args:
            workbook (Workbook): The Excel workbook object.
            project (Project): The project object.
        """
        sheet = workbook.active
        sheet.title = "Project"

        # Define column headers and types
        columns = [
            {"header": "Project Title", "type": "string", "width": 40},
            {"header": "Code", "type": "string", "width": 20},
            {"header": "Focal Person", "type": "string", "width": 20},
            {"header": "Email", "type": "string", "width": 25},
            {"header": "Project Description", "type": "string", "width": 50},
            {"header": "Organization", "type": "string", "width": 40},
            {"header": "Organization Type", "type": "string", "width": 40},
            {"header": "Cluster", "type": "string", "width": 50},
            {"header": "HRP Project Code", "type": "string", "width": 20},
            {"header": "Project Start Date", "type": "date", "width": 20},
            {"header": "Project End Date", "type": "date", "width": 30},
            {"header": "Project Budget", "type": "float", "width": 20},
            {"header": "Budget Received", "type": "float", "width": 20},
            {"header": "Budget Gap", "type": "float", "width": 20},
            {"header": "Project Budget Currency", "type": "string", "width": 20},
            {"header": "Project Donors", "type": "string", "width": 30},
            {"header": "Implementing Partners", "type": "string", "width": 30},
            {"header": "Programme Partners", "type": "string", "width": 30},
            {"header": "Status", "type": "string", "width": 10},
            {"header": "Activity Domain", "type": "string", "width": 50},
            {"header": "Activity Type", "type": "string", "width": 20},
            {"header": "Activity Detail", "type": "string", "width": 20},
            {"header": "Indicators", "type": "string", "width": 40},
            {"header": "admin1pcode", "type": "string", "width": 20},
            {"header": "admin1name", "type": "string", "width": 20},
            {"header": "region", "type": "string", "width": 20},
            {"header": "admin2pcode", "type": "string", "width": 20},
            {"header": "admin2name", "type": "string", "width": 20},
            {"header": "Zone/Ward", "type": "string", "width": 20},
        ]

        disaggregation_cols = []
        disaggregations = Disaggregation.objects.all()
        disaggregation_list = []

        for disaggregation in disaggregations:
            if disaggregation.name not in disaggregation_list:
                disaggregation_list.append(disaggregation.name)
                disaggregation_cols.append({"header": disaggregation.name, "type": "string", "width": 30})
            else:
                continue

        if disaggregations:
            for disaggregation_col in disaggregation_cols:
                columns.append(disaggregation_col)

        self.write_sheet_columns(sheet, columns)
        self.write_project_data_rows(sheet, project, columns, disaggregation_list)

    def write_sheet_columns(self, sheet, columns):
        """
        Write column headers and formatting for a sheet.

        Args:
            sheet (Worksheet): The worksheet object.
            columns (list): List of dictionaries containing column header, type, and width information.
        """
        for idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=idx, value=column["header"])
            cell.style = header_style

            column_letter = get_column_letter(idx)
            if column["type"] == "number":
                sheet.column_dimensions[column_letter].number_format = "General"
            elif column["type"] == "date":
                sheet.column_dimensions[column_letter].number_format = "mm-dd-yyyy"

            sheet.column_dimensions[column_letter].width = column["width"]

    def write_project_data_rows(self, sheet, project, columns, disaggregation_list):
        """
        Write project data rows to the sheet.

        Args:
            sheet (Worksheet): The worksheet object.
            project (Project): The project object.
        """
        rows = []
        try:
            plans = project.activityplan_set.all()
            for plan in plans:
                locations = plan.targetlocation_set.all()
                for location in locations:
                    # Create a dictionary to hold disaggregation data
                    disaggregation_data = {}

                    row = [
                        project.title,
                        project.code,
                        project.user.username if project.user else None,
                        project.user.email if project.user else None,
                        project.description,
                        project.user.profile.organization.code
                        if project.user and project.user.profile and project.user.profile.organization
                        else None,
                        project.user.profile.organization.type
                        if project.user and project.user.profile and project.user.profile.organization
                        else None,
                        ", ".join([clusters.code for clusters in project.clusters.all()]),
                        project.hrp_code,
                        project.start_date.astimezone(timezone.utc).replace(tzinfo=None)
                        if project.start_date
                        else None,
                        project.end_date.astimezone(timezone.utc).replace(tzinfo=None) if project.end_date else None,
                        project.budget,
                        project.budget_received,
                        project.budget_gap,
                        project.budget_currency.name if project.budget_currency else None,
                        ", ".join([donor.name for donor in project.donors.all()]),
                        ", ".join(
                            [implementing_partner.code for implementing_partner in project.implementing_partners.all()]
                        ),
                        ", ".join([programme_partner.code for programme_partner in project.programme_partners.all()]),
                        project.state,
                        plan.activity_domain.name if plan.activity_domain else None,
                        plan.activity_type.name if plan.activity_type else None,
                        plan.activity_detail.name if plan.activity_detail else None,
                        plan.indicator.name if plan.indicator else None,
                        location.province.code if location.province else None,
                        location.province.name if location.province else None,
                        location.province.region_name if location.province else None,
                        location.district.code if location.district else None,
                        location.district.name if location.district else None,
                        location.zone.name if location.zone else None,
                    ]
                    # Iterate through disaggregation locations and get disaggregation values
                    disaggregation_locations = location.disaggregationlocation_set.all()
                    disaggregation_location_list = {
                        disaggregation_location.disaggregation.name: disaggregation_location.target
                        for disaggregation_location in disaggregation_locations
                    }

                    # Update disaggregation_data with values from disaggregation_location_list
                    for disaggregation_entry in disaggregation_list:
                        if disaggregation_entry not in disaggregation_location_list:
                            disaggregation_data[disaggregation_entry] = None

                    disaggregation_location_list.update(disaggregation_data)

                    # Append disaggregation values to the row in the order of columns
                    for column in columns:
                        header = column["header"]
                        if header in disaggregation_location_list:
                            row.append(disaggregation_location_list[header])

                    # Add row to the list of rows
                    rows.append(row)

            for row_idx, row in enumerate(rows, start=2):
                for col_idx, value in enumerate(row, start=1):
                    try:
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                    except Exception as e:
                        print("Error:", e)

            # Correct syntax to freeze panes
            sheet.freeze_panes = "A2"

        except Exception as e:
            print("Error:", e)


class ProjectFilterExportView(View):
    def post(self, request, projectId):
        project = Project.objects.get(id=projectId)

        try:
            selectedData = json.loads(request.POST.get("exportData"))

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Project"
            # defining columns
            columns = [
                {"header": "Project Title", "type": "string", "width": 40},
                {"header": "Code", "type": "string", "width": 20},
                {"header": "Focal Person", "type": "string", "width": 20},
                {"header": "Email", "type": "string", "width": 25},
                {"header": "Organization", "type": "string", "width": 40},
                {"header": "Organization Type", "type": "string", "width": 40},
                {"header": "Project Description", "type": "string", "width": 50},
                {"header": "Cluster", "type": "string", "width": 50},
                {"header": "HRP project", "type": "string", "width": 50},
                {"header": "HRP Project Code", "type": "string", "width": 20},
                {"header": "Project Start Date", "type": "date", "width": 20},
                {"header": "Project End Date", "type": "date", "width": 30},
                {"header": "Project Budget", "type": "float", "width": 20},
                {"header": "Budget Received", "type": "float", "width": 20},
                {"header": "Budget Gap", "type": "float", "width": 20},
                {"header": "Project Budget Currency", "type": "string", "width": 20},
                {"header": "Project Donors", "type": "string", "width": 30},
                {"header": "Implementing Partners", "type": "string", "width": 30},
                {"header": "Programme Partners", "type": "string", "width": 30},
                {"header": "Status", "type": "string", "width": 10},
                {"header": "Activity Domain", "type": "string", "width": 50},
                {"header": "Activity Type", "type": "string", "width": 20},
                {"header": "Activity Detail", "type": "string", "width": 20},
                {"header": "Indicators", "type": "string", "width": 40},
                {"header": "Beneficiary", "type": "string", "width": 40},
                {"header": "Beneficiary category", "type": "string", "width": 40},
                {"header": "Activity description", "type": "string", "width": 40},
                {"header": "admin1pcode", "type": "string", "width": 20},
                {"header": "admin1name", "type": "string", "width": 20},
                {"header": "region", "type": "string", "width": 20},
                {"header": "admin2pcode", "type": "string", "width": 20},
                {"header": "admin2name", "type": "string", "width": 20},
                {"header": "classification", "type": "string", "width": 20},
                {"header": "Zone/Ward", "type": "string", "width": 20},
                {"header": "facility site type", "type": "string", "width": 20},
                {"header": "facility monitoring", "type": "string", "width": 20},
                {"header": "facility id", "type": "string", "width": 20},
                {"header": "facility name", "type": "string", "width": 20},
                {"header": "facility/site latitude", "type": "string", "width": 20},
                {"header": "facility/site longitude", "type": "string", "width": 20},
            ]

            disaggregation_cols = []
            disaggregations = Disaggregation.objects.all()
            disaggregation_list = []

            for disaggregation in disaggregations:
                if disaggregation.name not in disaggregation_list:
                    disaggregation_list.append(disaggregation.name)
                    disaggregation_cols.append({"header": disaggregation.name, "type": "string", "width": 30})
                else:
                    continue

            if disaggregations:
                for disaggregation_col in disaggregation_cols:
                    columns.append(disaggregation_col)

            for idx, column in enumerate(columns, start=1):
                cell = sheet.cell(row=1, column=idx, value=column["header"])
                cell.style = header_style

                column_letter = get_column_letter(idx)
                if column["type"] == "number":
                    sheet.column_dimensions[column_letter].number_format = "General"
                elif column["type"] == "date":
                    sheet.column_dimensions[column_letter].number_format = "mm-dd-yyyy"

                sheet.column_dimensions[column_letter].width = column["width"]
            # defining rows
            rows = []
            plans = project.activityplan_set.all()
            for plan in plans:
                locations = plan.targetlocation_set.all()
                for location in locations:
                    row = []
                    try:
                        row.append(
                            project.title if str(project.id) in selectedData["title"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.code if str(project.id) in selectedData["code"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.user.username if str(project.user.id) in selectedData["focal_point"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        if len(selectedData["focal_point"]) > 0:
                            row.append(
                                project.user.email if project.user else None,
                            )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        if len(selectedData["focal_point"]) > 0:
                            row.append(
                                project.user.profile.organization.code
                                if project.user and project.user.profile and project.user.profile.organization
                                else None,
                            )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        if len(selectedData["focal_point"]) > 0:
                            row.append(
                                project.user.profile.organization.type
                                if project.user and project.user.profile and project.user.profile.organization
                                else None,
                            )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.description if str(project.id) in selectedData["description"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            ", ".join(
                                [
                                    clusters.code
                                    for clusters in project.clusters.all()
                                    if str(clusters.id) in selectedData["clusters"]
                                ]
                            ),
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        if str(project.id) in selectedData["is_hrp_project"]:
                            row.append(
                                "yes" if project.is_hrp_project else None,
                            )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.has_hrp_code if str(project.id) in selectedData["hrp_code"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.start_date.astimezone(timezone.utc).replace(tzinfo=None)
                            if str(project.id) in selectedData["start_date"]
                            else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.end_date.astimezone(timezone.utc).replace(tzinfo=None)
                            if str(project.id) in selectedData["end_date"]
                            else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.budget if str(project.id) in selectedData["budget"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.budget_received if str(project.id) in selectedData["budget_received"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.budget_gap if str(project.id) in selectedData["budget_gap"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.budget_currency.name
                            if str(project.budget_currency.id) in selectedData["currency"]
                            else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            ", ".join(
                                [
                                    donor.name
                                    for donor in project.donors.all()
                                    if str(donor.id) in selectedData["donors"]
                                ]
                            ),
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            ", ".join(
                                [
                                    implementing_partner.code
                                    for implementing_partner in project.implementing_partners.all()
                                    if str(implementing_partner.id) in selectedData["implementing_partners"]
                                ]
                            ),
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            ", ".join(
                                [
                                    programme_partner.code
                                    for programme_partner in project.programme_partners.all()
                                    if str(programme_partner.id) in selectedData["programme_partners"]
                                ]
                            ),
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            project.state if str(project.id) in selectedData["state"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            plan.activity_domain.name
                            if str(plan.activity_domain.id) in selectedData["activity_domain"]
                            else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            plan.activity_type.name
                            if str(plan.activity_type.id) in selectedData["activity_type"]
                            else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            plan.activity_detail.name
                            if str(plan.activity_detail.id) in selectedData["activity_detail"]
                            else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            plan.indicator.name if str(plan.indicator.id) in selectedData["indicator"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            plan.beneficiary.name if str(plan.beneficiary.id) in selectedData["beneficiary"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            plan.beneficiary_category
                            if plan.beneficiary_category in selectedData["beneficiary_category"]
                            else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            plan.description if str(plan.id) in selectedData["activity_description"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.province.code if str(location.province.id) in selectedData["admin1name"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.province.name if str(location.province.id) in selectedData["admin1name"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.province.region_name
                            if str(location.province.id) in selectedData["admin1name"]
                            else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.district.code if str(location.district.id) in selectedData["admin2name"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.district.name if str(location.district.id) in selectedData["admin2name"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        if location.distict.classification:
                            row.append(
                                location.district.classification
                                if str(location.district.id) in selectedData["admin2name"]
                                else None,
                            )
                        else:
                            row.append("None,")
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.zone.name if str(location.province.id) in selectedData["admin1name"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.facility_site_type.name
                            if str(location.facility_site_type.id) in selectedData["facility_site_type"]
                            else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            str(location.facility_monitoring)
                            if str(location.id) in selectedData["facility_monitoring"]
                            else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.facility_id if str(location.id) in selectedData["facility_id"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.facility_name if str(location.id) in selectedData["facility_name"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.facility_lat if str(location.id) in selectedData["facility_latitude"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    try:
                        row.append(
                            location.facility_long if str(location.id) in selectedData["facility_longitude"] else None,
                        )
                    except Exception:
                        row.append(
                            " ",
                        )
                    target_location_disaggregation_list = []
                    try:
                        for item in selectedData["disaggregation"]:
                            target_location_disaggregation_list.append(item)
                    except Exception:
                        pass

                    disaggregation_data = {}
                    disaggregation_locations = location.disaggregationlocation_set.all()
                    disaggregation_location_list = {
                        disaggregation_location.disaggregation.name: disaggregation_location.target
                        for disaggregation_location in disaggregation_locations
                        if str(location.id) in target_location_disaggregation_list
                    }

                    # Update disaggregation_data with values from disaggregation_location_list
                    for disaggregation_entry in disaggregation_list:
                        if disaggregation_entry not in disaggregation_location_list:
                            disaggregation_data[disaggregation_entry] = None

                    disaggregation_location_list.update(disaggregation_data)

                    # Append disaggregation values to the row in the order of columns
                    for column in columns:
                        header = column["header"]
                        if header in disaggregation_location_list:
                            row.append(disaggregation_location_list[header])

                    rows.append(row)

            # row.extend(disaggregation_rows)
            for row_idx, row in enumerate(rows, start=2):
                for col_idx, value in enumerate(row, start=1):
                    try:
                        sheet.cell(row=row_idx, column=col_idx, value=value)
                    except Exception as e:
                        print("Error:", e)

            sheet.freeze_panes = sheet["A2"]
            excel_file = BytesIO()
            workbook.save(excel_file)
            excel_file.seek(0)

            response = {
                "file_url": "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
                + base64.b64encode(excel_file.read()).decode("utf-8"),
                "file_name": "export.xlsx",
            }

            return JsonResponse(response)
        except Exception as e:
            response = {"error": str(e)}
        return JsonResponse(response, status=500)


# Exporting the project in csv file
class ProjectExportCSV(View):
    def post(self, request, project_id):
        # Query the data you want to export
        project = Project.objects.get(id=project_id)

        # Create the HttpResponse object with CSV content type
        response = HttpResponse(project, content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=project.csv"

        # Create a CSV writer object
        writer = csv.writer(response)

        # Write headers to the CSV file
        headers = [
            "Project title",
            "Project code",
            "focal_point",
            "email",
            "Project Description",
            "Organization",
            "Organization Type",
            "Cluster",
            "HRP Project Code",
            "Project Start Date",
            "Project End Date",
            "Project Budget",
            "Budget Received",
            "Budget Gap",
            "Project Budget Currency",
            "Project Donors",
            "Implementing Partners",
            "Programme Partners",
            "Status",
            "Activity Domain",
            "Activity Type",
            "Activity Detail",
            "Indicators",
            "admin1pcode",
            "admin1name",
            "region",
            "admin2pcode",
            "admin2name",
            "Zone/Ward",
        ]
        # adding disaggreation headers
        disaggregation_cols = []
        disaggregations = Disaggregation.objects.all()
        disaggregation_list = []

        for disaggregation in disaggregations:
            if disaggregation.name not in disaggregation_list:
                disaggregation_list.append(disaggregation.name)
                disaggregation_cols.append(disaggregation.name)
            else:
                continue

        if disaggregations:
            for disaggregation_col in disaggregation_cols:
                headers.append(disaggregation_col)
        # writing headers to csv file
        writer.writerow(headers)

        rows = []
        # Write data to the CSV file
        plans = project.activityplan_set.all()
        for plan in plans:
            locations = plan.targetlocation_set.all()
            for location in locations:
                # Create a dictionary to hold disaggregation data
                row = [
                    project.title,
                    project.code,
                    project.user.username if project.user else None,
                    project.user.email if project.user else None,
                    project.description,
                    project.user.profile.organization.code
                    if project.user and project.user.profile and project.user.profile.organization
                    else None,
                    project.user.profile.organization.type
                    if project.user and project.user.profile and project.user.profile.organization
                    else None,
                    ", ".join([clusters.code for clusters in project.clusters.all()]),
                    project.hrp_code,
                    project.start_date.astimezone(timezone.utc).replace(tzinfo=None) if project.start_date else None,
                    project.end_date.astimezone(timezone.utc).replace(tzinfo=None) if project.end_date else None,
                    project.budget,
                    project.budget_received,
                    project.budget_gap,
                    project.budget_currency.name if project.budget_currency else None,
                    ", ".join([donor.name for donor in project.donors.all()]),
                    ", ".join(
                        [implementing_partner.code for implementing_partner in project.implementing_partners.all()]
                    ),
                    ", ".join([programme_partner.code for programme_partner in project.programme_partners.all()]),
                    project.state,
                    plan.activity_domain.name if plan.activity_domain else None,
                    plan.activity_type.name if plan.activity_type else None,
                    plan.activity_detail.name if plan.activity_detail else None,
                    plan.indicator.name if plan.indicator else None,
                    location.province.code if location.province else None,
                    location.province.name if location.province else None,
                    location.province.region_name if location.province else None,
                    location.district.code if location.district else None,
                    location.district.name if location.district else None,
                    location.zone.name if location.zone else None,
                ]
                disaggregation_data = {}
                disaggregation_locations = location.disaggregationlocation_set.all()
                disaggregation_location_list = {
                    disaggregation_location.disaggregation.name: disaggregation_location.target
                    for disaggregation_location in disaggregation_locations
                }
                # Update disaggregation_data with values from disaggregation_location_list
                for disaggregation_entry in disaggregation_list:
                    if disaggregation_entry not in disaggregation_location_list:
                        disaggregation_data[disaggregation_entry] = None

                disaggregation_location_list.update(disaggregation_data)
                for item in headers:
                    if item in disaggregation_location_list:
                        row.append(disaggregation_location_list[item])
                rows.append(row)
                writer.writerow(row)
        return response
