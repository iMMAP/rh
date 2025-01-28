import csv
import datetime

from dateutil.relativedelta import relativedelta
from django.db.models import Exists, OuterRef
from django.utils.timezone import now
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from project_reports.models import ActivityPlanReport, ProjectMonthlyReport, ResponseType
from rh.models import (
    Disaggregation,
    FacilitySiteType,
    GrantType,
    ImplementationModalityType,
    PackageType,
    RationSize,
    RationType,
    TransferCategory,
    TransferMechanismType,
    UnitType,
)

header_style = NamedStyle(name="header")
header_style.font = Font(bold=True)


def write_projects_reports_to_csv(monthly_progress_report, response):
    writer = csv.writer(response)
    columns = [
        "project_code",
        "report_id",
        "cluster_name",
        "focal_person_name",
        "focal_person_phone",
        "focal_person_email",
        "organization",
        "organization_type",
        "program_partner",
        "project_hrp-code",
        "project_title",
        "project_start_date",
        "project_end_date",
        "project_status",
        "response_types",
        "project_donor",
        "project_budget",
        "project_budget_currency",
        "report_month_number",
        "report_month",
        "report_year",
        "report_period",
        "implementing_partner",
        "admin0pcode",
        "admin0name",
        "region_name",
        "admin1pcode",
        "admin1name",
        "admin2pcode",
        "admin2name",
        "site_lat",
        "site_long",
        "facility_monitoring",
        "facility_site_type",
        "facility_site_name",
        "facility_site_id",
        "facility_site_lat",
        "facility_site_long",
        "non-hrp_beneficiary_code",
        "non-hrp_beneficiary_name",
        "hrp_beneficiary_code",
        "hrp_beneficiary_name",
        "beneficiary_status",
        "previously_assisted_by",
        "activity_domain_code",
        "activity_domain_name",
        "activity_type_code",
        "activity_type_name",
        "activity_detail_code",
        "activity_detail_name",
        "indicator_name",
        "units",
        "unit_type_name",
        "transfer_type_value",
        "implementation_modality_type_name",
        "transfer_mechanism_type_name",
        "package_type_name",
        "transfer_category_name",
        "grant_type",
        "currency",
        "updated_at",
        "created_at",
        "safe_spaces_for_women-girls",
    ]
    disaggregation_cols = []
    disaggregations = Disaggregation.objects.all()
    disaggregation_list = []

    for disaggregation in disaggregations:
        if disaggregation.name not in disaggregation_list:
            disaggregation_list.append(disaggregation.name)
            disaggregation_cols.append(disaggregation.name)
        else:
            continue

    if disaggregations:
        for disaggregation_col in disaggregation_cols:
            columns.append(disaggregation_col)

    # write csv columns
    writer.writerow(columns)

    try:
        for project_reports in monthly_progress_report:
            plan_reports = project_reports.activityplanreport_set.all()
            for plan_report in plan_reports:
                location_reports = plan_report.targetlocationreport_set.all()
                for location_report in location_reports:
                    # Create a dictionary to hold disaggregation data
                    disaggregation_data = {}
                    row = [
                        project_reports.project.code if project_reports.project.code else None,
                        project_reports.__str__(),
                        ", ".join(
                            [str(cluster.code) for cluster in plan_report.activity_plan.activity_domain.clusters.all()]
                        )
                        if project_reports.project.clusters
                        else None,
                        project_reports.project.user.first_name if project_reports.project.user.first_name else None,
                        project_reports.project.user.profile.phone if project_reports.project.user.profile else None,
                        project_reports.project.user.email if project_reports.project.user else None,
                        project_reports.project.user.profile.organization.code
                        if project_reports.project.user.profile
                        else None,
                        project_reports.project.user.profile.organization.type
                        if project_reports.project.user.profile
                        else None,
                        ", ".join([partner.name for partner in project_reports.project.programme_partners.all()])
                        if project_reports.project.programme_partners
                        else None,
                        project_reports.project.hrp_code if project_reports.project.hrp_code else None,
                        project_reports.project.title if project_reports.project.title else None,
                        project_reports.project.start_date.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                        if project_reports.project.start_date
                        else None,
                        project_reports.project.end_date.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                        if project_reports.project.end_date
                        else None,
                        project_reports.project.state if project_reports.project.state else None,
                        ", ".join(
                            [response_type.name for response_type in plan_report.response_types.all() if response_type]
                        )
                        if plan_report.response_types
                        else None,
                        ", ".join(str(donor.name) for donor in project_reports.project.donors.all())
                        if project_reports.project.donors
                        else None,
                        project_reports.project.budget if project_reports.project.budget else None,
                        project_reports.project.budget_currency.name
                        if project_reports.project.budget_currency
                        else None,
                        project_reports.from_date.month if project_reports.from_date else None,
                        project_reports.from_date.strftime("%B") if project_reports.from_date else None,
                        project_reports.from_date.strftime("%Y") if project_reports.from_date else None,
                        project_reports.from_date.strftime("%Y-%m-%d") if project_reports.from_date else None,
                        location_report.target_location.implementing_partner.name
                        if location_report.target_location.implementing_partner
                        else None,
                        location_report.target_location.country.code
                        if location_report.target_location.province
                        else None,
                        location_report.target_location.country.name
                        if location_report.target_location.province
                        else None,
                        location_report.target_location.province.region_name
                        if location_report.target_location.province
                        else None,
                        location_report.target_location.province.code
                        if location_report.target_location.province
                        else None,
                        location_report.target_location.province.name
                        if location_report.target_location.province
                        else None,
                        location_report.target_location.district.code
                        if location_report.target_location.district
                        else None,
                        location_report.target_location.district.name
                        if location_report.target_location.district
                        else None,
                        location_report.target_location.district.lat
                        if location_report.target_location.district
                        else None,
                        location_report.target_location.district.long
                        if location_report.target_location.district
                        else None,
                        "yes" if location_report.target_location.facility_monitoring else "No",
                        location_report.target_location.facility_site_type.name
                        if location_report.target_location.facility_site_type
                        else None,
                        location_report.target_location.facility_name
                        if location_report.target_location.facility_name
                        else None,
                        location_report.target_location.facility_id
                        if location_report.target_location.facility_id
                        else None,
                        location_report.target_location.facility_lat
                        if location_report.target_location.facility_lat
                        else None,
                        location_report.target_location.facility_long
                        if location_report.target_location.facility_long
                        else None,
                        plan_report.activity_plan.beneficiary.code if plan_report.activity_plan.beneficiary else None,
                        plan_report.activity_plan.beneficiary.name if plan_report.activity_plan.beneficiary else None,
                        plan_report.activity_plan.hrp_beneficiary.code
                        if plan_report.activity_plan.hrp_beneficiary
                        else None,
                        plan_report.activity_plan.hrp_beneficiary.name
                        if plan_report.activity_plan.hrp_beneficiary
                        else None,
                        location_report.beneficiary_status if location_report.beneficiary_status else None,
                        location_report.prev_assisted_by if location_report.prev_assisted_by else None,
                        plan_report.activity_plan.activity_domain.code
                        if plan_report.activity_plan.activity_domain
                        else None,
                        plan_report.activity_plan.activity_domain.name
                        if plan_report.activity_plan.activity_domain
                        else None,
                        plan_report.activity_plan.activity_type.code
                        if plan_report.activity_plan.activity_type
                        else None,
                        plan_report.activity_plan.activity_type.name
                        if plan_report.activity_plan.activity_type
                        else None,
                        plan_report.activity_plan.activity_detail.code
                        if plan_report.activity_plan.activity_detail
                        else None,
                        plan_report.activity_plan.activity_detail.name
                        if plan_report.activity_plan.activity_detail
                        else None,
                        plan_report.activity_plan.indicator.name if plan_report.activity_plan.indicator else None,
                        plan_report.units if plan_report.units else None,
                        plan_report.unit_type.name if plan_report.unit_type else None,
                        plan_report.no_of_transfers if plan_report.no_of_transfers else None,
                        plan_report.implement_modility_type.name if plan_report.implement_modility_type else None,
                        plan_report.transfer_mechanism_type.name if plan_report.transfer_mechanism_type else None,
                        plan_report.package_type.name if plan_report.package_type else None,
                        plan_report.transfer_category.name if plan_report.transfer_category else None,
                        plan_report.grant_type.name if plan_report.grant_type else None,
                        plan_report.currency.name if plan_report.currency else None,
                        project_reports.created_at.astimezone(datetime.timezone.utc).replace(tzinfo=None),
                        project_reports.updated_at.astimezone(datetime.timezone.utc).replace(tzinfo=None),
                        "Yes" if location_report.safe_space else None,
                    ]

                    # Iterate through disaggregation locations and get disaggregation values
                    disaggregation_locations = location_report.disaggregationlocationreport_set.all()
                    disaggregation_location_list = {
                        disaggregation_location.disaggregation.name: disaggregation_location.reached
                        for disaggregation_location in disaggregation_locations
                    }

                    # Update disaggregation_data with values from disaggregation_location_list
                    for disaggregation_entry in disaggregation_list:
                        if disaggregation_entry not in disaggregation_location_list:
                            disaggregation_data[disaggregation_entry] = None

                    disaggregation_location_list.update(disaggregation_data)

                    # Append disaggregation values to the row in the order of columns
                    for column_name in columns:
                        if column_name in disaggregation_location_list:
                            row.append(disaggregation_location_list[column_name])

                    writer.writerow(row)

    except Exception as e:
        print("Error:", e)


def write_import_report_template_sheet(workbook, monthly_report):
    sheet = workbook.active
    sheet.title = "Import Template"
    columns = [
        {"header": "project_code", "type": "string", "width": 40},
        {"header": "indicator", "type": "string", "width": 80},
        {"header": "activity_domain", "type": "string", "width": 40},
        {"header": "activity_type", "type": "string", "width": 80},
        {"header": "response_types", "type": "string", "width": 30},
        {"header": "implementing_partners", "type": "string", "width": 30},
        {"header": "package_type", "type": "string", "width": 30},
        {"header": "ration_type", "type": "string", "width": 30},
        {"header": "ration_size", "type": "string", "width": 30},
        {"header": "unit_type", "type": "string", "width": 30},
        {"header": "units/transfer value", "type": "string", "width": 30},
        {"header": "no_of_transfers", "type": "string", "width": 30},
        {"header": "grant_type", "type": "string", "width": 30},
        {"header": "transfer_category", "type": "string", "width": 30},
        {"header": "transfer_mechanism_type", "type": "string", "width": 30},
        {"header": "implement_modility_type", "type": "string", "width": 30},
        {"header": "beneficiary_status", "type": "string", "width": 30},
        {"header": "previously_assisted_by", "type": "string", "width": 50},
        {"header": "admin0name", "type": "string", "width": 30},
        {"header": "admin0pcode", "type": "string", "width": 30},
        {"header": "admin1pcode", "type": "string", "width": 30},
        {"header": "admin1name", "type": "string", "width": 30},
        {"header": "admin2pcode", "type": "string", "width": 30},
        {"header": "admin2name", "type": "string", "width": 30},
        {"header": "admin3pcode", "type": "string", "width": 30},
        {"header": "location_type", "type": "string", "width": 30},
        {"header": "facility_site_type", "type": "string", "width": 30},
        {"header": "facility_id", "type": "string", "width": 30},
        {"header": "facility_name", "type": "string", "width": 30},
        {"header": "facility_lat", "type": "string", "width": 30},
        {"header": "facility_long", "type": "string", "width": 30},
        {"header": "hrp_beneficiary", "type": "string", "width": 30},
        {"header": "with_safe_spaces", "type": "string", "width": 30},
    ]

    disaggregation_cols = []
    disaggregations = (
        Disaggregation.objects.filter(clusters__in=monthly_report.project.clusters.all()).order_by("-id").distinct()
    )
    for disaggregation in disaggregations:
        disaggregation_cols.append({"header": disaggregation.name, "type": "string", "width": 30})

    columns = columns + disaggregation_cols
    # write column headers in excel sheet
    for idx, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=idx, value=column["header"])
        cell.style = header_style
        # set column width
        column_letter = get_column_letter(idx)
        sheet.column_dimensions[column_letter].width = column["width"]

        if column["type"] == "number":
            sheet.column_dimensions[column_letter].number_format = "General"
        elif column["type"] == "date":
            sheet.column_dimensions[column_letter].number_format = "mm-dd-yyyy"

    sheet.column_dimensions[column_letter].width = column["width"]
    # write the rows with report data
    container_dictionary = {
        "beneficiary_status_list": ["Q", "New Beneficiary", "Existing Beneficiaries"],
        "reponseTypeList": ["E"],
        "package_type": ["G"],
        "ration_type": ["H"],
        "ration_size": ["I"],
        "unit_type": ["J"],
        "grant_type": ["M"],
        "im_modility_type": ["P"],
        "transfer_mc_type": ["O"],
        "transfer_category": ["N"],
        "safe_space": ["AG", "True", "False"],
        "facilitySiteTypeList": ["AA"],
    }
    plain_dictionary_lists = {
        "indicatorList": ["B"],
        "activityDomainList": ["C"],
        "activityTypeList": ["D"],
        "admin0nameList": ["S"],
        "admin0pcodeList": ["T"],
        "admin1pcodeList": ["U"],
        "admin1nameList": ["V"],
        "admin2pcodeList": ["W"],
        "admin2nameList": ["X"],
        "implementing_partner_list": ["F"],
        "hrp_beneficiary": ["AF"],
    }
    project = monthly_report.project
    container_dictionary["package_type"].extend(list(PackageType.objects.values_list("name", flat=True)))
    container_dictionary["unit_type"].extend(list(UnitType.objects.values_list("name", flat=True)))
    container_dictionary["grant_type"].extend(list(GrantType.objects.values_list("name", flat=True)))
    container_dictionary["im_modility_type"].extend(
        list(ImplementationModalityType.objects.values_list("name", flat=True))
    )
    container_dictionary["ration_type"].extend(list(RationType.objects.values_list("name", flat=True)))
    container_dictionary["ration_size"].extend(list(RationSize.objects.values_list("name", flat=True)))
    container_dictionary["transfer_mc_type"].extend(list(TransferMechanismType.objects.values_list("name", flat=True)))
    container_dictionary["transfer_category"].extend(list(TransferCategory.objects.values_list("name", flat=True)))

    facility = list(FacilitySiteType.objects.filter(cluster__code="health").values_list("name", flat=True))
    responseType = list(ResponseType.objects.values_list("name", flat=True))
    num_rows = 2

    cluster_code = []
    project_code = project.code
    for plan in project.activityplan_set.all():
        cluster_code.extend(plan.activity_domain.clusters.values_list("code", flat=True))
        for location in plan.targetlocation_set.all():
 
            plain_dictionary_lists["indicatorList"].append(str(plan.indicator.name))
            plain_dictionary_lists["activityDomainList"].append(str(plan.activity_domain.name))
            plain_dictionary_lists["activityTypeList"].append(str(plan.activity_type.name))
            plain_dictionary_lists["admin0pcodeList"].append(str(location.country.code))
            plain_dictionary_lists["admin0nameList"].append(str(location.country.name))
            plain_dictionary_lists["admin1nameList"].append(str(location.province.name))
            plain_dictionary_lists["admin1pcodeList"].append(str(location.province.code))
            plain_dictionary_lists["admin2pcodeList"].append(str(location.district.code))
            plain_dictionary_lists["admin2nameList"].append(str(location.district.name))
            plain_dictionary_lists["implementing_partner_list"].append(str(location.implementing_partner.code))
            plain_dictionary_lists["hrp_beneficiary"].append(str(plan.hrp_beneficiary))
            num_rows += 1

    container_dictionary["reponseTypeList"].extend(responseType)
    if "health" in cluster_code:
        container_dictionary["facilitySiteTypeList"].extend(facility)

    for key, value in container_dictionary.items():
        column = value[0]
        list_values = ",".join(value[1:])

        if len(list_values) < 255:
            dv = DataValidation(type="list", formula1='"{}"'.format(list_values))
            sheet.add_data_validation(dv)
            for row in range(2, num_rows):
                sheet[f"A{row}"] = project_code
                cell = sheet[f"{column}{row}"]
                sheet.add_data_validation(dv)
                dv.add(cell)

        # Case when the list values exceed 255 characters
        else:
            start_row = 2
            for i, item in enumerate(container_dictionary[key][1:]):
                sheet[f"{column}{start_row + i}"] = item

            # Define a named range for the list of values
            named_range = f"{column}_List"  # You can name the range based on the column or other criteria
            sheet.parent.create_named_range(
                named_range, sheet, f"{column}{start_row}:{column}{start_row + len(list_values) - 1}"
            )

            # Create a data validation that references the named range
            dv = DataValidation(type="list", formula1=f"={named_range}", showDropDown=True)
            sheet.add_data_validation(dv)
            for row in range(2, num_rows):
                cell = sheet[f"{column}{row}"]
                sheet.add_data_validation(dv)
                dv.add(cell)
    # display as plain list in the corresponding columns
    for key, value in plain_dictionary_lists.items():
        column = value[0]
        long_list_values = ",".join(value[1:])
        start_row = 2
        for i, item in enumerate(plain_dictionary_lists[key][1:]):
            sheet[f"{column}{start_row + i}"] = item

        # Define a named range for the list of values
        named_range = f"{column}_List"  # You can name the range based on the column or other criteria
        sheet.parent.create_named_range(
            named_range, sheet, f"{column}{start_row}:{column}{start_row + len(long_list_values) - 1}"
        )

        # Create a data validation that references the named range
        dv = DataValidation(type="list", formula1=f"={named_range}", showDropDown=True)
        sheet.add_data_validation(dv)
        for row in range(2, num_rows):
            cell = sheet[f"{column}{row}"]
            sheet.add_data_validation(dv)
            dv.add(cell)


def get_project_reporting_months(project):
    start_date = project.start_date
    end_date = project.end_date
    today = now()
    current_date = start_date
    months = []
    # Remove records that are beyond the project's start_date and end_date and without associated ActivityPlanReport
    ProjectMonthlyReport.objects.filter(project=project).exclude(
        from_date__gte=start_date, to_date__lte=end_date
    ).annotate(has_activityplanreport=Exists(ActivityPlanReport.objects.filter(monthly_report=OuterRef("pk")))).filter(
        has_activityplanreport=False
    ).delete()

    # Initialize variables
    current_date = start_date
    while current_date <= today:
        from_date = current_date.replace(day=1)
        first_day_next_month = current_date + relativedelta(months=1)
        to_date = first_day_next_month - relativedelta(days=1)
        if to_date < today:
            state = "pending"
        elif from_date <= today <= to_date:
            state = "todo"
        report, created = ProjectMonthlyReport.objects.get_or_create(
            project=project, from_date=from_date, to_date=to_date, defaults={"state": state}
        )
        if not created:
            # If the record exists, update its fields
            report.save()

        months.append({"from_date": from_date, "to_date": to_date, "state": state})

        current_date += relativedelta(months=1)
    return months


def write_projects_organization_to_csv(monthly_reports, response):
    writer = csv.writer(response)
    columns = ["organization name", "organization Acryname", "organization_type", "orgnaization clusters"]
    # write csv columns
    writer.writerow(columns)
    rows = []
    try:
        for project_report in monthly_reports:
            row = [
                project_report.project.organization.name,
                project_report.project.organization.code,
                project_report.project.organization.type,
                ",".join([str(cluster.code) for cluster in project_report.project.organization.clusters.all()]),
            ]
            if row not in rows:
                rows.append(row)
        writer.writerows(rows)

    except Exception as e:
        print("Error:", e)


def write_focal_persons_to_csv(monthly_reports, response):
    writer = csv.writer(response)
    columns = [
        "username",
        "first name",
        "last name",
        "email",
        "phone",
        "whatsapp",
        "skype",
        "organization",
        "position",
        "clusters",
        "country",
    ]
    # write csv columns
    writer.writerow(columns)
    rows = []
    try:
        for project_report in monthly_reports:
            row = [
                project_report.project.user.username,
                project_report.project.user.first_name if project_report.project.user.first_name else None,
                project_report.project.user.last_name if project_report.project.user.last_name else None,
                project_report.project.user.email if project_report.project.user.email else None,
                project_report.project.user.profile.phone if project_report.project.user.profile.phone else None,
                project_report.project.user.profile.whatsapp if project_report.project.user.profile.whatsapp else None,
                project_report.project.user.profile.skype if project_report.project.user.profile.skype else None,
                project_report.project.user.profile.organization.code
                if project_report.project.user.profile.organization
                else None,
                project_report.project.user.profile.position if project_report.project.user.profile.position else None,
                ",".join([str(cluster.code) for cluster in project_report.project.user.profile.clusters.all()]),
                project_report.project.user.profile.country if project_report.project.user.profile.country else None,
            ]
            if row not in rows:
                rows.append(row)
        writer.writerows(rows)

    except Exception as e:
        print("Error:", e)
