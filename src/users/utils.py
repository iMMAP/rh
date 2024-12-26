import datetime

from django.contrib.auth.models import Group, User
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter


def is_cluster_lead_of(user: User, cluster_code: str) -> bool:
    return user.groups.filter(name__in=f"{cluster_code.upper()}_CLUSTER_LEADS").exists()


def is_cluster_lead(user: User, clusters: list) -> bool:
    cluster_lead_groups = [f"{cluster.upper()}_CLUSTER_LEADS" for cluster in clusters]

    # user should be at least lead of one of the clusters
    if not user.groups.filter(name__in=cluster_lead_groups).exists():
        return False

    return True


# define header style
header_style = NamedStyle(name="header")
header_style.font = Font(bold=True)


def has_permission(user: User, user_obj: User, permission: str = "") -> bool:
    if user.is_superuser:
        return True

    if permission:
        if not user.has_perm(permission):
            return False

    if not (user.profile.organization == user_obj.profile.organization):
        return False

    return True


def assign_default_permissions_to_group(source_group_name: str, target_group: Group):
    try:
        source_group = Group.objects.get(name=source_group_name)
        source_group_permissions = list(source_group.permissions.all())

        target_group.permissions.add(*source_group_permissions)
    except Group.DoesNotExist:
        pass
    except Exception:
        pass


# export
def write_users_sheet(workbook, users):
    """
    Write the organization user sheet to the workbook
    Args:
        workbook (Workbook): The Excel workbook object.
        user (profile): The user object.
    """
    sheet = workbook.active
    sheet.title = "Organization Members"
    # Define column headers and types
    columns = [
        {"header": "username", "type": "string", "width": 40},
        {"header": "first name", "type": "string", "width": 40},
        {"header": "last name", "type": "string", "width": 40},
        {"header": "email", "type": "string", "width": 40},
        {"header": "phone", "type": "string", "width": 40},
        {"header": "whatsapp", "type": "string", "width": 40},
        {"header": "skype", "type": "string", "width": 40},
        {"header": "is cluster contact", "type": "string", "width": 40},
        {"header": "position", "type": "string", "width": 40},
        {"header": "organization", "type": "string", "width": 40},
        {"header": "organization type", "type": "string", "width": 40},
        {"header": "clusters", "type": "string", "width": 40},
        {"header": "country", "type": "string", "width": 40},
        {"header": "last login", "type": "string", "width": 40},
        {"header": "date joined", "type": "string", "width": 40},
    ]

    # write column headers in excel sheet
    for idx, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=idx, value=column["header"])
        cell.style = header_style

    column_letter = get_column_letter(idx)
    if column["type"] == "number":
        sheet.column_dimensions[column_letter].number_format = "General"
    elif column["type"] == "date":
        sheet.column_dimensions[column_letter].number_format = "mm-dd-yyyy"

    sheet.column_dimensions[column_letter].width = column["width"]
    # write the rows with report data

    rows = []

    try:
        for user in users:
            row = [
                user.username if user.username else None,
                user.first_name if user.first_name else None,
                user.last_name if user.last_name else None,
                user.email if user.email else None,
                user.profile.phone if user.profile.phone else None,
                user.profile.whatsapp if user.profile.whatsapp else None,
                user.profile.skype if user.profile.skype else None,
                "Yes" if user.profile.is_cluster_contact else "No",
                user.profile.position if user.profile.position else None,
                user.profile.organization.code if user.profile.organization else None,
                user.profile.organization.type if user.profile.organization else None,
                ", ".join([cluster.code for cluster in user.profile.clusters.all()])
                if user.profile.clusters.all()
                else None,
                user.profile.country.name if user.profile.country else None,
                user.last_login.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                if user.last_login
                else "No Sign-in",
                user.date_joined.astimezone(datetime.timezone.utc).replace(tzinfo=None) if user.date_joined else None,
            ]

            # Add row to the list of rows
            rows.append(row)

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                try:
                    sheet.cell(row=row_idx, column=col_idx, value=value)
                except Exception as e:
                    print("Error:", e)

        # Correct syntax to freeze panes
        sheet.freeze_panes = "A2"
    except Exception as e:
        print("Error:", e)
