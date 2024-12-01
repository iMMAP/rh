import base64
import csv
import datetime
import json
from io import BytesIO

from django.db.models import Prefetch
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter

from rh.utils import DateTimeEncoder

from ..filters import ProjectsFilter
from ..models import ActivityPlan, Disaggregation, DisaggregationLocation, Project, TargetLocation

#############################################
############### Export Views #################
#############################################

# Define the header style
header_style = NamedStyle(name="header")
header_style.font = Font(bold=True)


def project_export_excel_view(request, format):
    selected_projects_id = json.loads(request.body)
    user = request.user
    user_org = user.profile.organization
    project_state = ["in-progress", "completed", "archived"]
    project_queryset = Project.objects.select_related("organization", "budget_currency", "user").prefetch_related(
        "implementing_partners",
        "programme_partners",
        "donors",
        "clusters",
        Prefetch(
            "activityplan_set",
            queryset=ActivityPlan.objects.prefetch_related(
                Prefetch(
                    "targetlocation_set",
                    queryset=TargetLocation.objects.prefetch_related(
                        Prefetch(
                            "disaggregationlocation_set",
                            queryset=DisaggregationLocation.objects.select_related("disaggregation"),
                        )
                    ),
                )
            ),
        ),
    )

    # check the user permission
    if user.has_perm("rh.view_cluster_projects") or user.has_perm("rh.add_organization"):
        projects = project_queryset.filter(organization=user_org)
    elif (
        user.has_perm("rh.view_org_projects")
        or user.has_perm("rh.add_project")
        or user.has_perm("users.view_org_users")
    ):
        projects = project_queryset.filter(organization=user_org)
    else:
        return HttpResponseForbidden("Permission Denied !")

    if selected_projects_id:
        projects = projects.filter(id__in=selected_projects_id)

    # filter integration
    if request.GET:
        project_filter = ProjectsFilter(request.GET, request=request, queryset=projects)
        if project_filter.qs.exists():
            projects = project_filter.qs
        else:
            projects = project_filter.qs
    projects = projects.filter(state__in=project_state)

    try:
        # selectedData = json.loads(request.POST.get("exportData"))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Project"

        # form the filename with today date
        today_date = datetime.date.today()
        # define CSV file
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename='project.csv'"
        writer = csv.writer(response)

        # defining columns
        columns = [
            {"header": "Project Title", "type": "string", "width": 40},
            {"header": "Code", "type": "string", "width": 20},
            {"header": "Focal Person", "type": "string", "width": 20},
            {"header": "Email", "type": "string", "width": 25},
            {"header": "Organization", "type": "string", "width": 40},
            {"header": "Organization Type", "type": "string", "width": 40},
            {"header": "Status", "type": "string", "width": 10},
            {"header": "Cluster", "type": "string", "width": 50},
            {"header": "HRP project", "type": "string", "width": 50},
            {"header": "Project HRP Code", "type": "string", "width": 20},
            {"header": "Project Start Date", "type": "date", "width": 20},
            {"header": "Project End Date", "type": "date", "width": 30},
            {"header": "Project Budget", "type": "float", "width": 20},
            {"header": "Budget Received", "type": "float", "width": 20},
            {"header": "Budget Gap", "type": "float", "width": 20},
            {"header": "Project Budget Currency", "type": "string", "width": 20},
            {"header": "Project Donors", "type": "string", "width": 30},
            {"header": "Implementing Partners", "type": "string", "width": 30},
            {"header": "Programme Partners", "type": "string", "width": 30},
            {"header": "Activity Domain", "type": "string", "width": 50},
            {"header": "Description & Objective", "type": "string", "width": 50},
        ]
        for idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=idx, value=column["header"])
            cell.style = header_style

            column_letter = get_column_letter(idx)
            if column["type"] == "number":
                sheet.column_dimensions[column_letter].number_format = "General"
            elif column["type"] == "date":
                sheet.column_dimensions[column_letter].number_format = "mm-dd-yyyy"

            sheet.column_dimensions[column_letter].width = column["width"]

        # defining the csv header file
        if format == "csv":
            header_list = []
            for item in columns:
                header_list.append(item["header"])
            writer.writerow(header_list)

        # defining rows
        rows = []
        for project in projects:
            row = [
                project.title,
                project.code,
                project.user.username if project.user else None,
                project.user.email if project.user else None,
                project.user.profile.organization.code
                if project.user and project.user.profile and project.user.profile.organization
                else None,
                project.user.profile.organization.type
                if project.user and project.user.profile and project.user.profile.organization
                else None,
                project.state,
                ", ".join([clusters.code for clusters in project.clusters.all()]),
                "yes" if project.is_hrp_project == 1 else None,
                project.hrp_code if project.hrp_code else None,
                project.start_date.astimezone(datetime.timezone.utc).replace(tzinfo=None),
                project.end_date.astimezone(datetime.timezone.utc).replace(tzinfo=None),
                project.budget if project.budget else None,
                project.budget_received if project.budget_received else None,
                project.budget_gap if project.budget_gap else None,
                project.budget_currency.name if project.budget_currency else None,
                ", ".join([donor.name for donor in project.donors.all()]) if project.donors else None,
                ", ".join([implementing_partner.code for implementing_partner in project.implementing_partners.all()])
                if project.implementing_partners
                else None,
                ", ".join([programme_partner.code for programme_partner in project.programme_partners.all()])
                if project.programme_partners
                else None,
                ", ".join([activity_domain.name for activity_domain in project.activity_domains.all()])
                if project.activity_domains
                else None,
                project.description if project.description else None,
            ]

            # Add row to the list of rows
            rows.append(row)

            # write row for CSV file
            writer.writerow(row)

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                try:
                    sheet.cell(row=row_idx, column=col_idx, value=value)
                except Exception as e:
                    print("Error:", e)

        sheet.freeze_panes = sheet["A2"]
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        # form the filename with today date
        today_date = datetime.date.today()
        # check the requested format
        if format == "xls":
            response = {
                "file_url": "data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,"
                + base64.b64encode(excel_file.read()).decode("utf-8"),
                "file_name": f"{request.user.profile.organization}_projects_extracted_on_{today_date}.xlsx",
            }
            return JsonResponse(response)
        elif format == "csv":
            return response
        elif format == "json":
            # serialize the none serialze data and create dump data
            json_data = json.dumps(rows, cls=DateTimeEncoder, indent=4)

            response = HttpResponse(json_data, content_type="application/json")
            response["Content-Disposition"] = 'attachment; filename="projects.json"'
            return response

    except Exception as e:
        response = {"error": str(e)}
    return JsonResponse(response, status=500)


def single_project_export(request, pk, format):
    project = get_object_or_404(Project, pk=pk)
    try:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Project"
        # get today's date for filename
        today_date = datetime.date.today()
        # set response object for csv
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="project_extracted_on_{today_date}.csv"'
        writer = csv.writer(response)
        # defining columns
        columns = [
            {"header": "Project Title", "type": "string", "width": 40},
            {"header": "Code", "type": "string", "width": 20},
            {"header": "Focal Person", "type": "string", "width": 20},
            {"header": "Email", "type": "string", "width": 25},
            {"header": "Organization", "type": "string", "width": 40},
            {"header": "Organization Type", "type": "string", "width": 40},
            {"header": "Project Description", "type": "string", "width": 50},
            {"header": "Cluster", "type": "string", "width": 50},
            {"header": "HRP project", "type": "string", "width": 50},
            {"header": "Project HRP Code", "type": "string", "width": 20},
            {"header": "Project Start Date", "type": "date", "width": 20},
            {"header": "Project End Date", "type": "date", "width": 30},
            {"header": "Project Budget", "type": "float", "width": 20},
            {"header": "Budget Received", "type": "float", "width": 20},
            {"header": "Budget Gap", "type": "float", "width": 20},
            {"header": "Project Budget Currency", "type": "string", "width": 20},
            {"header": "Project Donors", "type": "string", "width": 30},
            {"header": "Implementing Partners", "type": "string", "width": 30},
            {"header": "Programme Partners", "type": "string", "width": 30},
            {"header": "Status", "type": "string", "width": 10},
            {"header": "Activity Domain", "type": "string", "width": 50},
            {"header": "Activity Type", "type": "string", "width": 20},
            {"header": "Activity Detail", "type": "string", "width": 20},
            {"header": "Indicators", "type": "string", "width": 40},
            {"header": "Beneficiary", "type": "string", "width": 40},
            {"header": "HRP Beneficiary", "type": "string", "width": 40},
            {"header": "Activity description", "type": "string", "width": 40},
            {"header": "Package Type", "type": "string", "width": 20},
            {"header": "Unit Type", "type": "string", "width": 20},
            {"header": "Units", "type": "string", "width": 20},
            {"header": "No_of_Transfer", "type": "string", "width": 20},
            {"header": "Grant Type", "type": "string", "width": 20},
            {"header": "Transfer Category", "type": "string", "width": 20},
            {"header": "Currency", "type": "string", "width": 20},
            {"header": "Transfer mechanism type", "type": "string", "width": 20},
            {"header": "implement modility type", "type": "string", "width": 20},
            {"header": "admin0code", "type": "string", "width": 20},
            {"header": "admin0name", "type": "string", "width": 20},
            {"header": "admin1pcode", "type": "string", "width": 20},
            {"header": "admin1name", "type": "string", "width": 20},
            {"header": "region", "type": "string", "width": 20},
            {"header": "admin2pcode", "type": "string", "width": 20},
            {"header": "admin2name", "type": "string", "width": 20},
            {"header": "NHS Code", "type": "string", "width": 20},
            {"header": "classification", "type": "string", "width": 20},
            {"header": "Zone/Ward", "type": "string", "width": 20},
            {"header": "facility site type", "type": "string", "width": 20},
            {"header": "facility monitoring", "type": "string", "width": 20},
            {"header": "facility name", "type": "string", "width": 20},
            {"header": "facility id", "type": "string", "width": 20},
            {"header": "facility/site latitude", "type": "string", "width": 20},
            {"header": "facility/site longitude", "type": "string", "width": 20},
        ]

        disaggregation_cols = []
        disaggregations = Disaggregation.objects.all()
        disaggregation_list = []

        for disaggregation in disaggregations:
            if disaggregation.name not in disaggregation_list:
                disaggregation_list.append(disaggregation.name)
                disaggregation_cols.append({"header": disaggregation.name, "type": "string", "width": 30})
            else:
                continue
        disaggregation_list.append("total")
        disaggregation_cols.append({"header": "total", "type": "string", "width": 20})
        if disaggregations:
            for disaggregation_col in disaggregation_cols:
                columns.append(disaggregation_col)

        for idx, column in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=idx, value=column["header"])
            cell.style = header_style

            column_letter = get_column_letter(idx)
            if column["type"] == "number":
                sheet.column_dimensions[column_letter].number_format = "General"
            elif column["type"] == "date":
                sheet.column_dimensions[column_letter].number_format = "mm-dd-yyyy"

            sheet.column_dimensions[column_letter].width = column["width"]

        # defining the csv header file
        if format == "csv":
            header_list = []
            for item in columns:
                header_list.append(item["header"])
            writer.writerow(header_list)

        # defining rows
        rows = []
        plans = project.activityplan_set.all()
        for plan in plans:
            locations = plan.targetlocation_set.all()
            for location in locations:
                # Create a dictionary to hold disaggregation data
                disaggregation_data = {}
                # check the selectedData if the field exist in the list then add to the row otherwise set it None
                row = [
                    project.title,
                    project.code,
                    project.user.username if project.user else None,
                    project.user.email if project.user else None,
                    project.user.profile.organization.code
                    if project.user and project.user.profile and project.user.profile.organization
                    else None,
                    project.user.profile.organization.type
                    if project.user and project.user.profile and project.user.profile.organization
                    else None,
                    project.description if project.description else None,
                    ", ".join([clusters.code for clusters in project.clusters.all()]),
                    "yes" if project.is_hrp_project == 1 else None,
                    project.hrp_code if project.hrp_code else None,
                    project.start_date.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                    if project.start_date
                    else None,
                    project.end_date.astimezone(datetime.timezone.utc).replace(tzinfo=None)
                    if project.end_date
                    else None,
                    project.budget if project.budget else None,
                    project.budget_received if project.budget_received else None,
                    project.budget_gap if project.budget_gap else None,
                    project.budget_currency.name if project.budget_currency else None,
                    ", ".join([donor.name for donor in project.donors.all()]) if project.donors else None,
                    ", ".join(
                        [implementing_partner.code for implementing_partner in project.implementing_partners.all()]
                    )
                    if project.implementing_partners
                    else None,
                    ", ".join([programme_partner.code for programme_partner in project.programme_partners.all()])
                    if project.programme_partners
                    else None,
                    project.state if project.state else None,
                    plan.activity_domain.name if plan.activity_domain else None,
                    plan.activity_type.name if plan.activity_type else None,
                    plan.activity_detail.name if plan.activity_detail else None,
                    plan.indicator.name if plan.indicator else None,
                    plan.beneficiary.name if plan.beneficiary else None,
                    plan.hrp_beneficiary.name if plan.hrp_beneficiary else None,
                    plan.description if plan.description else None,
                    plan.package_type.name if plan.package_type else None,
                    plan.unit_type.name if plan.unit_type else None,
                    plan.units if plan.units else None,
                    plan.no_of_transfers if plan.no_of_transfers else None,
                    plan.grant_type.name if plan.grant_type else None,
                    plan.transfer_category.name if plan.transfer_category else None,
                    plan.currency.name if plan.currency else None,
                    plan.transfer_mechanism_type.name if plan.transfer_mechanism_type else None,
                    plan.implement_modility_type.name if plan.implement_modility_type else None,
                    location.country.code if location.province else None,
                    location.country.name if location.province else None,
                    location.province.code if location.province else None,
                    location.province.name if location.province else None,
                    location.province.region_name if location.province else None,
                    location.district.code if location.district else None,
                    location.district.name if location.district else None,
                    location.nhs_code if location.nhs_code else None,
                    location.district.classification if location.district.classification else None,
                    location.zone.name if location.zone else None,
                    location.facility_site_type.name if location.facility_site_type else None,
                    "Yes" if location.facility_monitoring else None,
                    location.facility_name if location.facility_name else None,
                    location.facility_id if location.facility_id else None,
                    location.facility_lat if location.facility_lat else None,
                    location.facility_long if location.facility_long else None,
                ]

                # Iterate through disaggregation locations and get disaggregation values
                disaggregation_locations = location.disaggregationlocation_set.all()
                disaggregation_location_list = {
                    disaggregation_location.disaggregation.name: disaggregation_location.target
                    for disaggregation_location in disaggregation_locations
                }
                total_disaggregation = 0
                for value in disaggregation_location_list.values():
                    total_disaggregation += value
                disaggregation_location_list["total"] = total_disaggregation
                # Update disaggregation_data with values from disaggregation_location_list
                for disaggregation_entry in disaggregation_list:
                    if disaggregation_entry not in disaggregation_location_list:
                        disaggregation_data[disaggregation_entry] = None

                disaggregation_location_list.update(disaggregation_data)

                # Append disaggregation values to the row in the order of columns
                for column in columns:
                    header = column["header"]
                    if header in disaggregation_location_list:
                        row.append(disaggregation_location_list[header])

                # Add row to the list of rows
                rows.append(row)

                # write row for CSV file
                writer.writerow(row)

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                try:
                    sheet.cell(row=row_idx, column=col_idx, value=value)
                except Exception as e:
                    print("Error:", e)

        sheet.freeze_panes = sheet["A2"]
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        # form the filename with today date
        today_date = datetime.date.today()
        # check the requested format
        if format == "xlsx":
            response = HttpResponse(
                excel_file, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response["Content-Disposition"] = f'attachment; filename="project_extracted_on_{today_date}.xlsx"'
            return response
        elif format == "json":
            # serialize the none serialze data and create dump data
            json_data = json.dumps(rows, cls=DateTimeEncoder, indent=4)
            response = HttpResponse(json_data, content_type="application/json")
            response["Content-Disposition"] = f'attachment; filename="project_extracted_on_{today_date}.json"'
            return response
        elif format == "csv":
            return response
    except Exception as e:
        response = {"error": str(e)}
    return JsonResponse(response, status=500)
